#!/usr/bin/env python3
"""
Create SOUM-DECO-3-Template.xlsx
- Products tab: all 62 current products
- Orders tab: header only (15 columns incl. Variant)
- Stock tab: current stock CSV data
- Statistics tab: 5 sections with QUERY formulas
"""
import json
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ============================================================
#  LOAD DATA
# ============================================================
with open('/tmp/products.json', 'r', encoding='utf-8') as f:
    products = json.load(f)

with open('/tmp/stock.csv', 'r', encoding='utf-8') as f:
    stock_rows = list(csv.reader(f))

# ============================================================
#  STYLES
# ============================================================
HEADER_FONT = Font(name='Inter', bold=True, size=11, color='1C1815')
HEADER_FILL = PatternFill(start_color='E8E4DC', end_color='E8E4DC', fill_type='solid')
SECTION_FONT = Font(name='Inter', bold=True, size=12, color='9A7E3A')
SECTION_FILL = PatternFill(start_color='F1ECE3', end_color='F1ECE3', fill_type='solid')
TITLE_FONT = Font(name='Inter', bold=True, size=16, color='1C1815')
TITLE_FILL = PatternFill(start_color='FAF8F4', end_color='FAF8F4', fill_type='solid')
DATA_FONT = Font(name='Inter', size=10, color='2A2520')
THIN_BORDER = Border(
    left=Side(style='thin', color='D4CDBF'),
    right=Side(style='thin', color='D4CDBF'),
    top=Side(style='thin', color='D4CDBF'),
    bottom=Side(style='thin', color='D4CDBF'),
)

# ============================================================
#  CREATE WORKBOOK
# ============================================================
wb = Workbook()
wb.properties.creator = "SOUM DECO"

# ============================================================
#  PRODUCTS TAB
# ============================================================
ws_products = wb.active
ws_products.title = "Products"

PRODUCTS_COLS = ['id', 'name', 'description', 'category', 'price', 'image', 'images',
                 'featured', 'isSpecialOffer', 'variations', 'variants', 'stock',
                 'highlights', 'sortOrder', 'badge', 'oldPrice', 'quantityTiers']

# Header
for col_idx, col_name in enumerate(PRODUCTS_COLS, 1):
    cell = ws_products.cell(row=1, column=col_idx, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

# Data rows
for row_idx, product in enumerate(products, 2):
    for col_idx, col_name in enumerate(PRODUCTS_COLS, 1):
        value = product.get(col_name, '')
        # Convert booleans
        if col_name in ('featured', 'isSpecialOffer'):
            value = 'true' if (value is True or value == '1' or value == 'true') else 'false'
        # Convert None to empty
        if value is None:
            value = ''
        cell = ws_products.cell(row=row_idx, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

# Freeze header
ws_products.freeze_panes = 'A2'

# Column widths
col_widths = [15, 40, 30, 20, 10, 60, 60, 8, 12, 20, 40, 8, 30, 8, 15, 10, 30]
for i, w in enumerate(col_widths, 1):
    ws_products.column_dimensions[chr(64+i) if i <= 26 else 'A'].width = w
# Fix for columns beyond Z
from openpyxl.utils import get_column_letter
for i, w in enumerate(col_widths, 1):
    ws_products.column_dimensions[get_column_letter(i)].width = w

# ============================================================
#  ORDERS TAB (header only, 15 columns incl. Variant)
# ============================================================
ws_orders = wb.create_sheet("Orders")

ORDERS_HEADER = ['Date', 'Status', 'Product', 'Qty', 'Unit Price', 'Shipping',
                 'Total', 'Customer', 'Phone', 'Wilaya', 'Commune',
                 'Delivery', 'Company', 'Notes', 'Variant']

for col_idx, col_name in enumerate(ORDERS_HEADER, 1):
    cell = ws_orders.cell(row=1, column=col_idx, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

ws_orders.freeze_panes = 'A2'

order_widths = [20, 12, 50, 8, 12, 10, 12, 25, 15, 20, 20, 15, 15, 40, 30]
for i, w in enumerate(order_widths, 1):
    ws_orders.column_dimensions[get_column_letter(i)].width = w

# ============================================================
#  STOCK TAB
# ============================================================
ws_stock = wb.create_sheet("Stock")

# Header
ws_stock.cell(row=1, column=1, value='Product Name').font = HEADER_FONT
ws_stock.cell(row=1, column=1).fill = HEADER_FILL
ws_stock.cell(row=1, column=1).border = THIN_BORDER
ws_stock.cell(row=1, column=2, value='Stock Count').font = HEADER_FONT
ws_stock.cell(row=1, column=2).fill = HEADER_FILL
ws_stock.cell(row=1, column=2).border = THIN_BORDER

# Data (skip header row from CSV)
for row_idx, row in enumerate(stock_rows[1:], 2):
    if len(row) >= 2:
        name = row[0].strip()
        count = row[1].strip()
        if not name:
            continue
        # Try to convert count to number
        try:
            count_val = int(count) if count else ''
        except ValueError:
            count_val = count
        ws_stock.cell(row=row_idx, column=1, value=name).font = DATA_FONT
        ws_stock.cell(row=row_idx, column=1).border = THIN_BORDER
        ws_stock.cell(row=row_idx, column=2, value=count_val).font = DATA_FONT
        ws_stock.cell(row=row_idx, column=2).border = THIN_BORDER

ws_stock.freeze_panes = 'A2'
ws_stock.column_dimensions['A'].width = 60
ws_stock.column_dimensions['B'].width = 15

# Conditional formatting for stock (0=red, 1-3=yellow, >3=green)
from openpyxl.formatting.rule import CellIsRule
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

last_stock_row = len(stock_rows) + 1
ws_stock.conditional_formatting.add(f'B2:B{last_stock_row}',
    CellIsRule(operator='equal', formula=['0'], fill=red_fill))
ws_stock.conditional_formatting.add(f'B2:B{last_stock_row}',
    CellIsRule(operator='between', formula=['1', '3'], fill=yellow_fill))
ws_stock.conditional_formatting.add(f'B2:B{last_stock_row}',
    CellIsRule(operator='greaterThan', formula=['3'], fill=green_fill))

# ============================================================
#  STATISTICS TAB
# ============================================================
ws_stats = wb.create_sheet("Statistics")

ws_stats.column_dimensions['A'].width = 350
ws_stats.column_dimensions['B'].width = 100
ws_stats.column_dimensions['C'].width = 100
ws_stats.column_dimensions['D'].width = 150
ws_stats.column_dimensions['E'].width = 150

def set_section_title(row, text):
    cell = ws_stats.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for col in range(1, 6):
        ws_stats.cell(row=row, column=col).fill = SECTION_FILL
    ws_stats.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

def set_header_cell(row, col, text):
    cell = ws_stats.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

def set_formula(row, col, formula):
    cell = ws_stats.cell(row=row, column=col, value=formula)
    cell.font = DATA_FONT

# Title
title_cell = ws_stats.cell(row=1, column=1, value='📊 SOUM DECO — Tableau de bord')
title_cell.font = TITLE_FONT
title_cell.fill = TITLE_FILL
ws_stats.merge_cells('A1:E1')
for col in range(1, 6):
    ws_stats.cell(row=1, column=col).fill = TITLE_FILL

# Section 1: Summary
set_section_title(3, '📦 Résumé')
ws_stats.cell(row=4, column=1, value='Total Commandes').font = Font(name='Inter', bold=True, size=10)
set_formula(4, 2, '=COUNTA(Orders!C2:C)')
ws_stats.cell(row=4, column=2).number_format = '0'

ws_stats.cell(row=5, column=1, value="Chiffre d'Affaires (DZD)").font = Font(name='Inter', bold=True, size=10)
set_formula(5, 2, '=SUM(Orders!G2:G)')
ws_stats.cell(row=5, column=2).number_format = '#,##0'

ws_stats.cell(row=6, column=1, value='Panier Moyen (DZD)').font = Font(name='Inter', bold=True, size=10)
set_formula(6, 2, '=IF(COUNTA(Orders!C2:C)>0, SUM(Orders!G2:G)/COUNTA(Orders!C2:C), 0)')
ws_stats.cell(row=6, column=2).number_format = '#,##0'

# Section 2: Top 10 Products
set_section_title(8, '🏆 Top 10 Produits')
set_header_cell(9, 1, 'Produit')
set_header_cell(9, 2, 'Commandes')
set_header_cell(9, 3, 'CA (DZD)')
set_formula(10, 1, '=QUERY(Orders!C2:G, "SELECT C, COUNT(C), SUM(G) WHERE C IS NOT NULL GROUP BY C ORDER BY COUNT(C) DESC LIMIT 10", 1)')

# Section 3: Top 10 Wilayas
set_section_title(22, '📍 Top 10 Wilayas')
set_header_cell(23, 1, 'Wilaya')
set_header_cell(23, 2, 'Commandes')
set_header_cell(23, 3, 'CA (DZD)')
set_formula(24, 1, '=QUERY(Orders!J2:G, "SELECT J, COUNT(J), SUM(G) WHERE J IS NOT NULL GROUP BY J ORDER BY COUNT(J) DESC LIMIT 10", 1)')

# Section 4: Order Status Breakdown
set_section_title(36, '📋 Statut des Commandes')
set_header_cell(37, 1, 'Statut')
set_header_cell(37, 2, 'Nombre')
set_formula(38, 1, '=QUERY(Orders!B2:B, "SELECT B, COUNT(B) WHERE B IS NOT NULL GROUP BY B ORDER BY COUNT(B) DESC LIMIT 10", 1)')

# Section 5: Top 5 Wilayas by Revenue
set_section_title(50, '💰 Top 5 Wilayas par CA')
set_header_cell(51, 1, 'Wilaya')
set_header_cell(51, 2, 'CA (DZD)')
set_formula(52, 1, '=QUERY(Orders!J2:G, "SELECT J, SUM(G) WHERE J IS NOT NULL GROUP BY J ORDER BY SUM(G) DESC LIMIT 5", 1)')

ws_stats.freeze_panes = 'A2'

# ============================================================
#  SAVE
# ============================================================
output_path = '/home/z/my-project/download/SOUM-DECO-3-Template.xlsx'
wb.save(output_path)
print(f"✅ Saved: {output_path}")
print(f"   Products: {len(products)} rows")
print(f"   Orders: header only (15 columns incl. Variant)")
print(f"   Stock: {len(stock_rows)-1} entries")
print(f"   Statistics: 5 sections with QUERY formulas")
