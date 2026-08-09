#!/usr/bin/env python3
"""
Build the Soum Deco Google Sheet template (.xlsx) — beautiful, guided, colored.
3 tabs: Products (29 products pre-filled), Orders, Stock.
No Shipping tab.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---- Load the 29 selected products ----
with open('/tmp/selected_products.json') as f:
    products = json.load(f)

# ---- Column structure (matches apps-script.gs PRODUCTS_COLS exactly) ----
PRODUCTS_COLS = [
    'id', 'name', 'description', 'category', 'price', 'image', 'images',
    'featured', 'isSpecialOffer', 'variations', 'variants', 'stock',
    'highlights', 'sortOrder', 'badge', 'oldPrice', 'quantityTiers'
]

# ============================================================
#  COLOR PALETTE (matches the website: warm cream + brass + charcoal)
# ============================================================
BRAND_CHARCOAL = "2A2520"      # dark text / header bg
BRAND_BRASS    = "9A7E3A"      # accent / section title bg
BRASS_LIGHT    = "F5EFE2"      # light brass for guidance row
CREAM          = "FAF8F4"      # warm ivory background
SAND           = "F1ECE3"      # subtle backgrounds
CLAY           = "D4CDBF"      # borders
WHITE          = "FFFFFF"
GRAY_TEXT      = "6B6358"      # muted text
INK            = "1C1815"      # strong text

# Status colors (Orders tab)
STATUS_NEW       = "3080FF"    # blue
STATUS_CONFIRMED = "016630"    # green
STATUS_SHIPPED   = "FFD700"    # yellow
STATUS_DELIVERED = "2F7D5B"    # dark green
STATUS_CANCELLED = "E40014"    # red

# Stock colors
STOCK_OUT   = "E40014"   # red (0)
STOCK_LOW   = "FFD700"   # yellow (1-3)
STOCK_OK    = "2F7D5B"   # green (>3)

# Arabic guidance row (row 2) — bilingual hints
ARABIC_GUIDANCE = {
    'id': '🆔 المعرّف — يجب أن يكون فريداً',
    'name': '🏷️ اسم المنتج — يظهر للعملاء',
    'description': '📝 الوصف — التفاصيل الكاملة',
    'category': '📂 الفئة — تجمع المنتجات في أقسام',
    'price': '💰 السعر بالدينار (فارغ = عند الطلب)',
    'image': '🖼️ رابط الصورة الأولى (الغلاف)',
    'images': '🗂️ كل الصور (مفصولة بـ ~~~)',
    'featured': '⭐ مميّز؟ اكتب true أو false',
    'isSpecialOffer': '🎁 عرض خاص؟ اكتب true أو false',
    'variations': '🔄 التنوعات (قديم — اتركه فارغ)',
    'variants': '🎨 المقاسات والألوان',
    'stock': '📦 المخزون (فارغ = غير محدود)',
    'highlights': '✨ أبرز المميزات (سطر لكل ميزة)',
    'sortOrder': '🔢 الترتيب (رقم أصغر = يظهر أولاً)',
    'badge': '🎖️ الشارة (مثال: عرض خاص، جديد)',
    'oldPrice': '💸 السعر القديم (اختياري — يظهر مشطوب)',
    'quantityTiers': '🎉 عروض الكمية (مثال: 2:desk:500,3:both:1000)',
}

# ============================================================
#  STYLES
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=CLAY),
    right=Side(style='thin', color=CLAY),
    top=Side(style='thin', color=CLAY),
    bottom=Side(style='thin', color=CLAY),
)

medium_border_bottom = Border(
    left=Side(style='thin', color=CLAY),
    right=Side(style='thin', color=CLAY),
    top=Side(style='thin', color=CLAY),
    bottom=Side(style='medium', color=BRAND_BRASS),
)

# Header row style (row 1) — dark charcoal bg, white bold text
header_font = Font(name='Calibri', bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=BRAND_CHARCOAL, end_color=BRAND_CHARCOAL, fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Guidance row style (row 2) — light brass bg, italic muted text
guidance_font = Font(name='Calibri', bold=False, size=10, color=GRAY_TEXT, italic=True)
guidance_fill = PatternFill(start_color=BRASS_LIGHT, end_color=BRASS_LIGHT, fill_type='solid')
guidance_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data row style — bold dark text on cream bg
data_font = Font(name='Calibri', bold=True, size=10, color=INK)
data_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Alternating row fill (zebra stripes)
zebra_fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

# ---- Create workbook ----
wb = openpyxl.Workbook()

# ============================================================
#  PRODUCTS TAB
# ============================================================
ws = wb.active
ws.title = 'Products'
ws.sheet_properties.tabColor = BRAND_BRASS  # colored tab

# Row 1: English headers
for col_idx, col_name in enumerate(PRODUCTS_COLS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    cell.border = medium_border_bottom

# Row 2: Arabic guidance
for col_idx, col_name in enumerate(PRODUCTS_COLS, 1):
    cell = ws.cell(row=2, column=col_idx, value=ARABIC_GUIDANCE.get(col_name, ''))
    cell.font = guidance_font; cell.fill = guidance_fill; cell.alignment = guidance_align
    cell.border = thin_border

# Rows 3+: product data with zebra stripes
for row_idx, p in enumerate(products, start=3):
    values = {
        'id': p['id'], 'name': p['name'], 'description': p.get('description', ''),
        'category': p.get('category', ''), 'price': p.get('price'),
        'image': p.get('image', ''), 'images': p.get('images', ''),
        'featured': 'true' if p.get('featured') else 'false', 'isSpecialOffer': 'false',
        'variations': '', 'variants': '', 'stock': '', 'highlights': '',
        'sortOrder': row_idx - 2, 'badge': '', 'oldPrice': '', 'quantityTiers': '',
    }
    is_zebra = (row_idx % 2 == 1)  # odd rows get cream
    for col_idx, col_name in enumerate(PRODUCTS_COLS, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=values[col_name])
        cell.font = data_font
        cell.alignment = center_align if col_name in ('price', 'sortOrder', 'stock', 'featured', 'isSpecialOffer') else data_align
        cell.border = thin_border
        cell.fill = zebra_fill if is_zebra else white_fill
    # Highlight featured products with a subtle brass tint on the featured column
    if p.get('featured'):
        fc = ws.cell(row=row_idx, column=PRODUCTS_COLS.index('featured') + 1)
        fc.fill = PatternFill(start_color="E8D9B0", end_color="E8D9B0", fill_type='solid')
        fc.font = Font(name='Calibri', bold=True, size=10, color=BRAND_BRASS)

# Column widths
col_widths = {'id': 18, 'name': 32, 'description': 45, 'category': 18, 'price': 12, 'image': 50, 'images': 60, 'featured': 11, 'isSpecialOffer': 13, 'variations': 14, 'variants': 18, 'stock': 11, 'highlights': 25, 'sortOrder': 9, 'badge': 14, 'oldPrice': 11, 'quantityTiers': 22}
for col_idx, col_name in enumerate(PRODUCTS_COLS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)
ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 42
for row_idx in range(3, 3 + len(products)):
    ws.row_dimensions[row_idx].height = 55
ws.freeze_panes = 'A3'  # freeze header + guidance rows

# Dropdowns for featured + isSpecialOffer columns (true/false) — no typing needed
from openpyxl.worksheet.datavalidation import DataValidation
featured_col_letter = get_column_letter(PRODUCTS_COLS.index('featured') + 1)
special_col_letter = get_column_letter(PRODUCTS_COLS.index('isSpecialOffer') + 1)
bool_dv = DataValidation(
    type="list",
    formula1='"true,false"',
    allow_blank=True,
    showDropDown=False,  # False = SHOW the dropdown arrow (openpyxl quirk)
)
bool_dv.prompt = 'Click to choose: true or false'
bool_dv.promptTitle = 'Yes / No'
ws.add_data_validation(bool_dv)
bool_dv.add(f'{featured_col_letter}3:{featured_col_letter}1000')
bool_dv.add(f'{special_col_letter}3:{special_col_letter}1000')

# ============================================================
#  ORDERS TAB
# ============================================================
ws_o = wb.create_sheet('Orders')
ws_o.sheet_properties.tabColor = STATUS_NEW  # blue tab

orders_headers = ['Date', 'Status', 'Product', 'Qty', 'Unit Price', 'Shipping', 'Total', 'Customer', 'Phone', 'Wilaya', 'Commune', 'Delivery', 'Company', 'Notes']
orders_guidance = [
    '📅 تلقائي — وقت الطلب',
    '🔵 الحالة: New → Confirmed → Shipped → Delivered',
    '🛍️ اسم المنتج',
    '🔢 الكمية',
    '💰 سعر الوحدة',
    '🚚 سعر التوصيل',
    '💵 المجموع الكلي',
    '👤 اسم الزبون',
    '📞 رقم الهاتف',
    '📍 الولاية',
    '🏘️ البلدية',
    '🏠 طريقة التوصيل',
    '🏢 شركة التوصيل',
    '📝 ملاحظات',
]

for col_idx, h in enumerate(orders_headers, 1):
    cell = ws_o.cell(row=1, column=col_idx, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    cell.border = medium_border_bottom
for col_idx, g in enumerate(orders_guidance, 1):
    cell = ws_o.cell(row=2, column=col_idx, value=g)
    cell.font = guidance_font; cell.fill = guidance_fill; cell.alignment = guidance_align
    cell.border = thin_border

orders_widths = [18, 14, 30, 6, 11, 9, 11, 20, 14, 14, 16, 16, 18, 25]
for col_idx, w in enumerate(orders_widths, 1):
    ws_o.column_dimensions[get_column_letter(col_idx)].width = w
ws_o.row_dimensions[1].height = 32
ws_o.row_dimensions[2].height = 42
ws_o.freeze_panes = 'A3'

# Status conditional formatting (vivid colors)
status_col = get_column_letter(2)
status_range = f'{status_col}3:{status_col}1000'
ws_o.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"New"'], fill=PatternFill(start_color=STATUS_NEW, end_color=STATUS_NEW, fill_type='solid'), font=Font(bold=True, color=WHITE, size=11)))
ws_o.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Confirmed"'], fill=PatternFill(start_color=STATUS_CONFIRMED, end_color=STATUS_CONFIRMED, fill_type='solid'), font=Font(bold=True, color=WHITE, size=11)))
ws_o.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Shipped"'], fill=PatternFill(start_color=STATUS_SHIPPED, end_color=STATUS_SHIPPED, fill_type='solid'), font=Font(bold=True, color=INK, size=11)))
ws_o.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Delivered"'], fill=PatternFill(start_color=STATUS_DELIVERED, end_color=STATUS_DELIVERED, fill_type='solid'), font=Font(bold=True, color=WHITE, size=11)))
ws_o.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Cancelled"'], fill=PatternFill(start_color=STATUS_CANCELLED, end_color=STATUS_CANCELLED, fill_type='solid'), font=Font(bold=True, color=WHITE, size=11)))

# Status dropdown (data validation) — user picks from list, no typing needed
status_dv = DataValidation(
    type="list",
    formula1='"New,Confirmed,Shipped,Delivered,Cancelled"',
    allow_blank=True,
    showDropDown=False,  # False = SHOW the dropdown arrow (openpyxl quirk)
)
status_dv.error = 'Please select a status from the dropdown'
status_dv.errorTitle = 'Invalid status'
status_dv.prompt = 'Click to choose: New → Confirmed → Shipped → Delivered (or Cancelled)'
status_dv.promptTitle = 'Order Status'
ws_o.add_data_validation(status_dv)
status_dv.add(status_range)

# ============================================================
#  STOCK TAB
# ============================================================
ws_s = wb.create_sheet('Stock')
ws_s.sheet_properties.tabColor = STOCK_OK  # green tab

stock_headers = ['Product Name', 'Stock Count']
stock_guidance = [
    '🏷️ نفس اسم المنتج في تبويب Products (انسخه والصقه هنا)',
    '📦 اكتب رقم: 10، 3، 0 (فارغ = غير محدود)',
]

for col_idx, h in enumerate(stock_headers, 1):
    cell = ws_s.cell(row=1, column=col_idx, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    cell.border = medium_border_bottom
for col_idx, g in enumerate(stock_guidance, 1):
    cell = ws_s.cell(row=2, column=col_idx, value=g)
    cell.font = guidance_font; cell.fill = guidance_fill; cell.alignment = guidance_align
    cell.border = thin_border

# Pre-fill stock tab with all 29 product names
for row_idx, p in enumerate(products, start=3):
    is_zebra = (row_idx % 2 == 1)
    cell_name = ws_s.cell(row=row_idx, column=1, value=p['name'])
    cell_name.font = data_font; cell_name.alignment = data_align; cell_name.border = thin_border
    cell_name.fill = zebra_fill if is_zebra else white_fill
    cell_count = ws_s.cell(row=row_idx, column=2, value='')
    cell_count.font = data_font; cell_count.alignment = center_align; cell_count.border = thin_border
    cell_count.fill = zebra_fill if is_zebra else white_fill

ws_s.column_dimensions['A'].width = 42
ws_s.column_dimensions['B'].width = 20
ws_s.row_dimensions[1].height = 32
ws_s.row_dimensions[2].height = 42
ws_s.freeze_panes = 'A3'

# Stock conditional formatting
stock_range = 'B3:B1000'
ws_s.conditional_formatting.add(stock_range, CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color=STOCK_OUT, end_color=STOCK_OUT, fill_type='solid'), font=Font(bold=True, color=WHITE, size=11)))
ws_s.conditional_formatting.add(stock_range, CellIsRule(operator='between', formula=['1', '3'], fill=PatternFill(start_color=STOCK_LOW, end_color=STOCK_LOW, fill_type='solid'), font=Font(bold=True, color=INK, size=11)))
ws_s.conditional_formatting.add(stock_range, CellIsRule(operator='greaterThan', formula=['3'], fill=PatternFill(start_color=STOCK_OK, end_color=STOCK_OK, fill_type='solid'), font=Font(bold=True, color=WHITE, size=11)))

# ============================================================
#  STATISTICS TAB — professional dashboard with live formulas
# ============================================================
#  All calculations use COUNTIF/SUMIF/COUNTIFS formulas that read
#  from the Orders tab. When you add/edit/delete orders, the
#  statistics update AUTOMATICALLY — no manual refresh needed.
# ============================================================
ws_st = wb.create_sheet('Statistics')
ws_st.sheet_properties.tabColor = BRAND_BRASS  # brass tab

# Helper styles for stat cards
stat_title_font = Font(name='Calibri', bold=True, size=14, color=BRAND_CHARCOAL)
stat_value_font = Font(name='Calibri', bold=True, size=24, color=BRAND_BRASS)
stat_label_font = Font(name='Calibri', size=10, color=GRAY_TEXT)
stat_card_fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type='solid')
stat_card_border = Border(
    left=Side(style='medium', color=BRAND_BRASS),
    right=Side(style='thin', color=CLAY),
    top=Side(style='thin', color=CLAY),
    bottom=Side(style='thin', color=CLAY),
)

# Section header style
section_font = Font(name='Calibri', bold=True, size=13, color=WHITE)
section_fill = PatternFill(start_color=BRAND_BRASS, end_color=BRAND_BRASS, fill_type='solid')
section_align = Alignment(horizontal='center', vertical='center')

# Table header style
table_header_font = Font(name='Calibri', bold=True, size=11, color=WHITE)
table_header_fill = PatternFill(start_color=BRAND_CHARCOAL, end_color=BRAND_CHARCOAL, fill_type='solid')

# --- Title row ---
ws_st.merge_cells('A1:F1')
title_cell = ws_st.cell(row=1, column=1, value='📊 لوحة الإحصائيات · Tableau de bord')
title_cell.font = Font(name='Calibri', bold=True, size=18, color=WHITE)
title_cell.fill = PatternFill(start_color=BRAND_CHARCOAL, end_color=BRAND_CHARCOAL, fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws_st.row_dimensions[1].height = 40

# --- Section 1: Key Metrics (cards) ---
# Row 3: section header
ws_st.merge_cells('A3:F3')
s1 = ws_st.cell(row=3, column=1, value='🔑 المؤشرات الرئيسية · Indicateurs clés')
s1.font = section_font; s1.fill = section_fill; s1.alignment = section_align
ws_st.row_dimensions[3].height = 28

# Row 4-5: 4 stat cards (2 rows × 4 cols layout using A:C and D:F)
# Card 1: Total Orders
ws_st.cell(row=4, column=1, value='📦 إجمالي الطلبات').font = stat_label_font
ws_st.cell(row=4, column=1).fill = stat_card_fill
ws_st.cell(row=5, column=1, value='=COUNTA(Orders!C3:C10000)').font = stat_value_font
ws_st.cell(row=5, column=1).fill = stat_card_fill
ws_st.cell(row=5, column=1).alignment = Alignment(horizontal='center', vertical='center')
for r in [4, 5]:
    for c in [1]:
        ws_st.cell(row=r, column=c).border = stat_card_border

# Card 2: Total Revenue (Grand Total sum, excluding Cancelled)
ws_st.cell(row=4, column=2, value='💰 إجمالي المبيعات').font = stat_label_font
ws_st.cell(row=4, column=2).fill = stat_card_fill
ws_st.cell(row=5, column=2, value='=SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"<>Cancelled")').font = stat_value_font
ws_st.cell(row=5, column=2).fill = stat_card_fill
ws_st.cell(row=5, column=2).number_format = '#,##0 "دج"'
ws_st.cell(row=5, column=2).alignment = Alignment(horizontal='center', vertical='center')
for r in [4, 5]:
    for c in [2]:
        ws_st.cell(row=r, column=c).border = stat_card_border

# Card 3: Total Shipping Revenue
ws_st.cell(row=4, column=3, value='🚚 إجمالي التوصيل').font = stat_label_font
ws_st.cell(row=4, column=3).fill = stat_card_fill
ws_st.cell(row=5, column=3, value='=SUMIFS(Orders!F3:F10000,Orders!B3:B10000,"<>Cancelled")').font = stat_value_font
ws_st.cell(row=5, column=3).fill = stat_card_fill
ws_st.cell(row=5, column=3).number_format = '#,##0 "دج"'
ws_st.cell(row=5, column=3).alignment = Alignment(horizontal='center', vertical='center')
for r in [4, 5]:
    for c in [3]:
        ws_st.cell(row=r, column=c).border = stat_card_border

# Card 4: Average Order Value
ws_st.cell(row=4, column=4, value='📈 متوسط قيمة الطلب').font = stat_label_font
ws_st.cell(row=4, column=4).fill = stat_card_fill
ws_st.cell(row=5, column=4, value='=IFERROR(AVERAGEIFS(Orders!G3:G10000,Orders!B3:B10000,"<>Cancelled"),0)').font = stat_value_font
ws_st.cell(row=5, column=4).fill = stat_card_fill
ws_st.cell(row=5, column=4).number_format = '#,##0 "دج"'
ws_st.cell(row=5, column=4).alignment = Alignment(horizontal='center', vertical='center')
for r in [4, 5]:
    for c in [4]:
        ws_st.cell(row=r, column=c).border = stat_card_border

# Card 5: Pending Orders (New)
ws_st.cell(row=4, column=5, value='🆕 طلبات جديدة').font = stat_label_font
ws_st.cell(row=4, column=5).fill = stat_card_fill
ws_st.cell(row=5, column=5, value='=COUNTIF(Orders!B3:B10000,"New")').font = stat_value_font
ws_st.cell(row=5, column=5).fill = stat_card_fill
ws_st.cell(row=5, column=5).alignment = Alignment(horizontal='center', vertical='center')
for r in [4, 5]:
    for c in [5]:
        ws_st.cell(row=r, column=c).border = stat_card_border

# Card 6: Delivered Orders
ws_st.cell(row=4, column=6, value='✅ طلبات مكتملة').font = stat_label_font
ws_st.cell(row=4, column=6).fill = stat_card_fill
ws_st.cell(row=5, column=6, value='=COUNTIF(Orders!B3:B10000,"Delivered")').font = stat_value_font
ws_st.cell(row=5, column=6).fill = stat_card_fill
ws_st.cell(row=5, column=6).alignment = Alignment(horizontal='center', vertical='center')
for r in [4, 5]:
    for c in [6]:
        ws_st.cell(row=r, column=c).border = stat_card_border

ws_st.row_dimensions[4].height = 20
ws_st.row_dimensions[5].height = 40

# --- Section 2: Order Status Breakdown ---
ws_st.merge_cells('A7:F7')
s2 = ws_st.cell(row=7, column=1, value='📋 تفصيل الطلبات · Statut des commandes')
s2.font = section_font; s2.fill = section_fill; s2.alignment = section_align
ws_st.row_dimensions[7].height = 28

# Table headers
status_headers = ['الحالة · Statut', 'العدد · Nombre', 'النسبة · %', 'المبلغ · Montant']
for col_idx, h in enumerate(status_headers, 1):
    cell = ws_st.cell(row=8, column=col_idx, value=h)
    cell.font = table_header_font; cell.fill = table_header_fill
    cell.alignment = center_align; cell.border = thin_border

# Status rows with formulas
status_data = [
    ('🆕 New', '=COUNTIF(Orders!B3:B10000,"New")', '=IFERROR(B9/SUM($B$9:$B$13),0)', '=SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"New")', STATUS_NEW),
    ('✅ Confirmed', '=COUNTIF(Orders!B3:B10000,"Confirmed")', '=IFERROR(B10/SUM($B$9:$B$13),0)', '=SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"Confirmed")', STATUS_CONFIRMED),
    ('🚚 Shipped', '=COUNTIF(Orders!B3:B10000,"Shipped")', '=IFERROR(B11/SUM($B$9:$B$13),0)', '=SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"Shipped")', STATUS_SHIPPED),
    ('📦 Delivered', '=COUNTIF(Orders!B3:B10000,"Delivered")', '=IFERROR(B12/SUM($B$9:$B$13),0)', '=SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"Delivered")', STATUS_DELIVERED),
    ('❌ Cancelled', '=COUNTIF(Orders!B3:B10000,"Cancelled")', '=IFERROR(B13/SUM($B$9:$B$13),0)', '=SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"Cancelled")', STATUS_CANCELLED),
]
for row_idx, (label, count_formula, pct_formula, amount_formula, color) in enumerate(status_data, start=9):
    # Status label with color
    c1 = ws_st.cell(row=row_idx, column=1, value=label)
    c1.font = Font(name='Calibri', bold=True, size=11, color=WHITE if color != STATUS_SHIPPED else INK)
    c1.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    c1.alignment = center_align; c1.border = thin_border
    # Count
    c2 = ws_st.cell(row=row_idx, column=2, value=count_formula)
    c2.font = data_font; c2.alignment = center_align; c2.border = thin_border
    # Percentage
    c3 = ws_st.cell(row=row_idx, column=3, value=pct_formula)
    c3.font = data_font; c3.alignment = center_align; c3.border = thin_border
    c3.number_format = '0.0%'
    # Amount
    c4 = ws_st.cell(row=row_idx, column=4, value=amount_formula)
    c4.font = data_font; c4.alignment = center_align; c4.border = thin_border
    c4.number_format = '#,##0 "دج"'

# --- Section 3: Top Products ---
ws_st.merge_cells('A15:F15')
s3 = ws_st.cell(row=15, column=1, value='🏆 المنتجات الأكثر مبيعاً · Top produits')
s3.font = section_font; s3.fill = section_fill; s3.alignment = section_align
ws_st.row_dimensions[15].height = 28

# Table headers
product_headers = ['المنتج · Produit', 'الكمية المباعة · Qty', 'الإيراد · Revenu', 'النسبة · %']
for col_idx, h in enumerate(product_headers, 1):
    cell = ws_st.cell(row=16, column=col_idx, value=h)
    cell.font = table_header_font; cell.fill = table_header_fill
    cell.alignment = center_align; cell.border = thin_border

# Top 10 products — formulas read from Orders tab column C (Product)
# Uses COUNTIF for quantity and SUMIFS for revenue
# Note: these show ALL orders (including cancelled) for product analysis
for i in range(10):
    row_idx = 17 + i
    rank = i + 1
    # Product name: we can't dynamically extract top products with formulas alone.
    # Instead, we list the top 10 product names from the Products tab and count their orders.
    if i < len(products):
        product_name = products[i]['name']
        # Count how many times this product appears in Orders (column C)
        c1 = ws_st.cell(row=row_idx, column=1, value=f'{rank}. {product_name}')
        c1.font = data_font; c1.alignment = data_align; c1.border = thin_border
        # Quantity sold: count occurrences in Orders column C
        c2 = ws_st.cell(row=row_idx, column=2, value=f'=COUNTIF(Orders!C3:C10000,"*{product_name}*")')
        c2.font = data_font; c2.alignment = center_align; c2.border = thin_border
        # Revenue: sum of Total (column G) where Product contains this name
        c3 = ws_st.cell(row=row_idx, column=3, value=f'=SUMIFS(Orders!G3:G10000,Orders!C3:C10000,"*{product_name}*")')
        c3.font = data_font; c3.alignment = center_align; c3.border = thin_border
        c3.number_format = '#,##0 "دج"'
        # Percentage of total revenue
        c4 = ws_st.cell(row=row_idx, column=4, value=f'=IFERROR(C{row_idx}/SUMIFS(Orders!G3:G10000,Orders!B3:B10000,"<>Cancelled"),0)')
        c4.font = data_font; c4.alignment = center_align; c4.border = thin_border
        c4.number_format = '0.0%'

# --- Section 4: Shipping Companies ---
start_row = 28
ws_st.merge_cells(f'A{start_row}:F{start_row}')
s4 = ws_st.cell(row=start_row, column=1, value='🚚 شركات التوصيل · Sociétés de livraison')
s4.font = section_font; s4.fill = section_fill; s4.alignment = section_align
ws_st.row_dimensions[start_row].height = 28

# Table headers
company_headers = ['الشركة · Société', 'الطلبات · Commandes', 'الإيراد · Revenu']
for col_idx, h in enumerate(company_headers, 1):
    cell = ws_st.cell(row=start_row + 1, column=col_idx, value=h)
    cell.font = table_header_font; cell.fill = table_header_fill
    cell.alignment = center_align; cell.border = thin_border

# Company rows
companies = ['ZR Express', 'Ecom Delivery']
for i, company in enumerate(companies):
    row_idx = start_row + 2 + i
    c1 = ws_st.cell(row=row_idx, column=1, value=company)
    c1.font = data_font; c1.alignment = data_align; c1.border = thin_border
    c2 = ws_st.cell(row=row_idx, column=2, value=f'=COUNTIF(Orders!M3:M10000,"{company}")')
    c2.font = data_font; c2.alignment = center_align; c2.border = thin_border
    c3 = ws_st.cell(row=row_idx, column=3, value=f'=SUMIFS(Orders!G3:G10000,Orders!M3:M10000,"{company}")')
    c3.font = data_font; c3.alignment = center_align; c3.border = thin_border
    c3.number_format = '#,##0 "دج"'

# --- Section 5: Delivery Types ---
start_row2 = start_row + 5
ws_st.merge_cells(f'A{start_row2}:F{start_row2}')
s5 = ws_st.cell(row=start_row2, column=1, value='🏠 طرق التوصيل · Modes de livraison')
s5.font = section_font; s5.fill = section_fill; s5.alignment = section_align
ws_st.row_dimensions[start_row2].height = 28

delivery_headers = ['الطريقة · Mode', 'الطلبات · Commandes', 'الإيراد · Revenu']
for col_idx, h in enumerate(delivery_headers, 1):
    cell = ws_st.cell(row=start_row2 + 1, column=col_idx, value=h)
    cell.font = table_header_font; cell.fill = table_header_fill
    cell.alignment = center_align; cell.border = thin_border

delivery_types = [
    ('مكتب التوصيل (Stop Desk)', '=COUNTIF(Orders!L3:L10000,"مكتب التوصيل")'),
    ('توصيل للمنزل (Home)', '=COUNTIF(Orders!L3:L10000,"توصيل للمنزل")'),
]
for i, (label, count_formula) in enumerate(delivery_types):
    row_idx = start_row2 + 2 + i
    c1 = ws_st.cell(row=row_idx, column=1, value=label)
    c1.font = data_font; c1.alignment = data_align; c1.border = thin_border
    c2 = ws_st.cell(row=row_idx, column=2, value=count_formula)
    c2.font = data_font; c2.alignment = center_align; c2.border = thin_border
    c3 = ws_st.cell(row=row_idx, column=3, value=f'=SUMIFS(Orders!G3:G10000,Orders!L3:L10000,"{label.split(" (")[0]}")')
    c3.font = data_font; c3.alignment = center_align; c3.border = thin_border
    c3.number_format = '#,##0 "دج"'

# --- Section 6: Wilaya Analysis ---
start_row3 = start_row2 + 5
ws_st.merge_cells(f'A{start_row3}:F{start_row3}')
s6 = ws_st.cell(row=start_row3, column=1, value='📍 الولايات الأكثر طلباً · Wilayas les plus actives')
s6.font = section_font; s6.fill = section_fill; s6.alignment = section_align
ws_st.row_dimensions[start_row3].height = 28

wilaya_headers = ['الولاية · Wilaya', 'الطلبات · Commandes', 'الإيراد · Revenu']
for col_idx, h in enumerate(wilaya_headers, 1):
    cell = ws_st.cell(row=start_row3 + 1, column=col_idx, value=h)
    cell.font = table_header_font; cell.fill = table_header_fill
    cell.alignment = center_align; cell.border = thin_border

# Note: Wilaya analysis requires the admin to manually list the wilayas they want to track.
# The formulas will count orders for each wilaya listed in column A.
# Pre-fill with common wilayas (Alger, Oran, Constantine, etc.)
common_wilayas = [
    '16 - Alger', '31 - Oran', '25 - Constantine', '23 - Annaba',
    '06 - Béjaïa', '15 - Tizi Ouzou', '19 - Sétif', '09 - Blida',
]
for i, wilaya in enumerate(common_wilayas):
    row_idx = start_row3 + 2 + i
    c1 = ws_st.cell(row=row_idx, column=1, value=wilaya)
    c1.font = data_font; c1.alignment = data_align; c1.border = thin_border
    c2 = ws_st.cell(row=row_idx, column=2, value=f'=COUNTIF(Orders!J3:J10000,"{wilaya}")')
    c2.font = data_font; c2.alignment = center_align; c2.border = thin_border
    c3 = ws_st.cell(row=row_idx, column=3, value=f'=SUMIFS(Orders!G3:G10000,Orders!J3:J10000,"{wilaya}")')
    c3.font = data_font; c3.alignment = center_align; c3.border = thin_border
    c3.number_format = '#,##0 "دج"'

# --- Column widths for Statistics tab ---
ws_st.column_dimensions['A'].width = 35
ws_st.column_dimensions['B'].width = 18
ws_st.column_dimensions['C'].width = 18
ws_st.column_dimensions['D'].width = 18
ws_st.column_dimensions['E'].width = 18
ws_st.column_dimensions['F'].width = 18

# Hide gridlines for a cleaner dashboard look
ws_st.sheet_view.showGridLines = False

# ---- Save ----
output_path = '/home/z/my-project/download/Soum-Deco-Sheet-Template.xlsx'
wb.save(output_path)
print(f'✓ Saved: {output_path}')
print(f'  Products tab: {len(products)} products (with zebra stripes, brass-tinted featured column)')
print(f'  Orders tab: empty (vivid status colors: blue/green/yellow/dark-green/red)')
print(f'  Stock tab: {len(products)} product names pre-filled (red/yellow/green stock colors)')
print(f'  Statistics tab: 6 sections with live formulas (key metrics, status, top products, companies, delivery, wilayas)')
print(f'  No Shipping tab')
