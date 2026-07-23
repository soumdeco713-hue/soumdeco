"""Final sheet with vivid status colors and bold text."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

NAVY = '0A1E3A'
GOLD = 'D4AF37'
WHITE = 'FFFFFF'
LIGHT_GOLD = 'FFF8E1'
GRAY_TEXT = '6B7B95'

# Vivid status colors
COLOR_NEW = 'E3F2FD'        # light blue
COLOR_CONFIRMED = 'C8E6C9'  # light green
COLOR_SHIPPED = 'FFF9C4'    # light yellow
COLOR_DELIVERED = 'E8F5E9'  # vivid green
COLOR_CANCELLED = 'FFCDD2'  # light red

COLOR_INSTOCK = 'C8E6C9'    # green
COLOR_OUTSTOCK = 'FFCDD2'   # red
COLOR_LOWSTOCK = 'FFF9C4'   # yellow

header_font = Font(name='Arial', bold=True, size=13, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
gold_header_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type='solid')
gold_header_font = Font(name='Arial', bold=True, size=13, color=NAVY)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D8DCE4'),
    right=Side(style='thin', color='D8DCE4'),
    top=Side(style='thin', color='D8DCE4'),
    bottom=Side(style='thin', color='D8DCE4'),
)
note_font = Font(name='Arial', size=10, italic=True, color=GRAY_TEXT)
note_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type='solid')
bold_data = Font(name='Arial', size=11, color=NAVY, bold=True)

def style_header(ws, num_cols, use_gold=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = gold_header_font if use_gold else header_font
        cell.fill = gold_header_fill if use_gold else header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'A2'

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = note_font
    cell.fill = note_fill
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

def apply_bold_to_data(ws, num_cols, row_start=3, row_end=5000):
    """Apply bold font to all data cells."""
    for row in range(row_start, row_end + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = bold_data

# ================================================================
# TAB 1: PRODUCTS
# ================================================================
ws_p = wb.active
ws_p.title = 'Products'

p_headers = ['id','name','description','category','price','oldPrice','image','images','featured','isSpecialOffer','variants','stock','quantityTiers','badge','sortOrder']
p_widths = [18, 30, 50, 20, 12, 12, 40, 50, 10, 12, 35, 10, 30, 15, 10]
p_notes = ['المعرّف','اسم المنتج','الوصف','الفئة','السعر (دج)','السعر القديم','رابط الصورة','كل الصور','مميّز: true/false','عرض خاص: true/false','المقاسات والألوان','المخزون (فارغ=غير محدود)','عروض الكمية','الشارة','الترتيب']

for col, header in enumerate(p_headers, 1):
    ws_p.cell(row=1, column=col, value=header)

style_header(ws_p, len(p_headers))
set_widths(ws_p, p_widths)

dv1 = DataValidation(type='list', formula1='"true,false"', allow_blank=True)
ws_p.add_data_validation(dv1)
dv1.add('I2:I5000')
dv2 = DataValidation(type='list', formula1='"true,false"', allow_blank=True)
ws_p.add_data_validation(dv2)
dv2.add('J2:J5000')

for col, note in enumerate(p_notes, 1):
    add_note(ws_p, 2, col, note)
ws_p.row_dimensions[2].height = 25

# ================================================================
# TAB 2: ORDERS — with vivid status colors
# ================================================================
ws_o = wb.create_sheet('Orders')

o_headers = ['Date','Status','Product','Qty','Unit Price','Shipping','Total','Customer','Phone','Wilaya','Commune','Delivery','Company','Notes']
o_widths = [18, 16, 30, 8, 12, 10, 12, 20, 15, 15, 20, 15, 15, 30]
o_notes = ['تلقائي','New → Confirmed → Shipped → Delivered','','','','','','','','','','','','']

for col, header in enumerate(o_headers, 1):
    ws_o.cell(row=1, column=col, value=header)

style_header(ws_o, len(o_headers), use_gold=True)
set_widths(ws_o, o_widths)

dv3 = DataValidation(type='list', formula1='"New,Confirmed,Shipped,Delivered,Cancelled"', allow_blank=True)
ws_o.add_data_validation(dv3)
dv3.add('B2:B5000')

for col, note in enumerate(o_notes, 1):
    if note:
        add_note(ws_o, 2, col, note)
ws_o.row_dimensions[2].height = 25

# Conditional formatting — vivid status colors on column B (Status)
ws_o.conditional_formatting.add('B2:B5000', CellIsRule(operator='equal', formula=['"New"'], fill=PatternFill(start_color=COLOR_NEW, end_color=COLOR_NEW, fill_type='solid'), font=Font(name='Arial', bold=True, size=11, color='0D47A1')))
ws_o.conditional_formatting.add('B2:B5000', CellIsRule(operator='equal', formula=['"Confirmed"'], fill=PatternFill(start_color=COLOR_CONFIRMED, end_color=COLOR_CONFIRMED, fill_type='solid'), font=Font(name='Arial', bold=True, size=11, color='1B5E20')))
ws_o.conditional_formatting.add('B2:B5000', CellIsRule(operator='equal', formula=['"Shipped"'], fill=PatternFill(start_color=COLOR_SHIPPED, end_color=COLOR_SHIPPED, fill_type='solid'), font=Font(name='Arial', bold=True, size=11, color='F57F17')))
ws_o.conditional_formatting.add('B2:B5000', CellIsRule(operator='equal', formula=['"Delivered"'], fill=PatternFill(start_color=COLOR_DELIVERED, end_color=COLOR_DELIVERED, fill_type='solid'), font=Font(name='Arial', bold=True, size=11, color='1B5E20')))
ws_o.conditional_formatting.add('B2:B5000', CellIsRule(operator='equal', formula=['"Cancelled"'], fill=PatternFill(start_color=COLOR_CANCELLED, end_color=COLOR_CANCELLED, fill_type='solid'), font=Font(name='Arial', bold=True, size=11, color='B71C1C')))

# ================================================================
# TAB 3: STOCK — with vivid stock colors
# ================================================================
ws_s = wb.create_sheet('Stock')

s_headers = ['Product Name', 'Stock Count']
s_widths = [40, 15]
s_notes = ['نفس اسم المنتج في تبويب Products', 'اكتب رقم: 10، 3، 0 (فارغ=غير محدود)']

for col, header in enumerate(s_headers, 1):
    ws_s.cell(row=1, column=col, value=header)

style_header(ws_s, len(s_headers))
set_widths(ws_s, s_widths)

for col, note in enumerate(s_notes, 1):
    add_note(ws_s, 2, col, note)
ws_s.row_dimensions[2].height = 25

# Conditional formatting on column B (Stock Count):
# 0 = red, 1-3 = yellow, >3 = green
ws_s.conditional_formatting.add('B2:B5000', CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color=COLOR_OUTSTOCK, end_color=COLOR_OUTSTOCK, fill_type='solid'), font=Font(name='Arial', bold=True, size=12, color='B71C1C')))
ws_s.conditional_formatting.add('B2:B5000', CellIsRule(operator='between', formula=['1','3'], fill=PatternFill(start_color=COLOR_LOWSTOCK, end_color=COLOR_LOWSTOCK, fill_type='solid'), font=Font(name='Arial', bold=True, size=12, color='E65100')))
ws_s.conditional_formatting.add('B2:B5000', CellIsRule(operator='greaterThan', formula=['3'], fill=PatternFill(start_color=COLOR_INSTOCK, end_color=COLOR_INSTOCK, fill_type='solid'), font=Font(name='Arial', bold=True, size=12, color='1B5E20')))

# ===== SAVE =====
output = '/home/z/my-project/download/El-Miizaan-Sheet-Template.xlsx'
wb.save(output)
print(f'Saved: {output}')
